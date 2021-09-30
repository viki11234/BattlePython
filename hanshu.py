import random

random_guard_list = list()
random_elite_list = list()
from openpyxl import load_workbook
from openpyxl.styles import Font,Color

random_levelset = [[2, 5, 8], [3, 7], [3, 6, 8], [1, 4, 7]]

wb = load_workbook('S3关卡随机池.xlsx')
ws = wb['Sheet1']
for i in range(4, 12):
    random_guard_list.append(ws["b" + str(i)].value)
for i in range(4, 9):
    random_elite_list.append(ws["e" + str(i)].value)

font = Font(color="FF0000")

ws = wb.create_sheet("newsheet")
i = 1
m = 0
while i < 31:
    pp = random.randint(1,100)<20

    if i % 10 == 1:
        ws["a" + str(m + 1)] = i // 10 + 1
    ws["b" + str(m + 1)] = i % 10

    if i % 10 == 0:
        ws["b" + str(m + 1)] = 10

    ws["c" + str(m + 1)] = 1
    index = random.randint(0, len(random_guard_list) - 1)
    ws["d" + str(m + 1)].value = (random_guard_list[index])

    ws["c" + str(m + 2)] = 2
    index2 = random.randint(0, len(random_guard_list) - 1)
    while index == index2:
        index2 = random.randint(0, len(random_guard_list) - 1)
    ws["d" + str(m + 2)].value = (random_guard_list[index2])

    ws["c" + str(m + 3)] = 3
    ws["d" + str(m + 3)].value = (random_elite_list[random.randint(0, len(random_elite_list) - 1)])

    if pp :
        ws["d" + str(m + 1)].font = font
        ws["d" + str(m + 2)].font = font
        ws["d" + str(m + 3)].font = font

    m = m + 3
    if random.randint(1, 100) < 20:
        ws["d" + str(m + 1)].value = (random_elite_list[random.randint(0, len(random_elite_list) - 1)])
        ws["c" + str(m + 1)] = 4
        if pp :
            ws["d" + str(m + 1)].font = font
        m = m + 1

    i += 1
wb.save('S3关卡随机池.xlsx')
